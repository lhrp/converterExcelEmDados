from openpyxl import load_workbook
from tkinter import filedialog
from tkinter import messagebox
from ttkbootstrap.constants import *
from win11toast import notify
import customtkinter as ctk
import mysql.connector
import os
import pandas as pd
import pyodbc
import sqlite3
import ttkbootstrap as ttk



telaPrincipal = ttk.Window()
telaPrincipal.title("Importador de Excel")
#Tamanho Padrão
width = 400
height = 350
#Tamanho Mínimo
# minWidth = 400
# minHeight = 350
#Resolução atual do windows
largura_tela = telaPrincipal.winfo_screenwidth()
altura_tela = telaPrincipal.winfo_screenheight()
#Necessário para centralizar
posX = largura_tela / 2 - width / 2
posY = altura_tela / 2 - height / 2
#Tamanho Máximo
#maxWidth = largura_tela
#maxHeight = altura_tela

#Definindo os tamanhos máximos e mínimos da tela
#telaPrincipal.maxsize(width=maxWidth, height=maxHeight)
#telaPrincipal.minsize(width=minWidth, height=minHeight)
telaPrincipal.resizable(width=False, height=False)
#Forçando o foco na tela aberta
telaPrincipal.focus_force()
#Definindo o tema padrão da aplicação
style = ttk.Style(theme="superhero")
#Centralizando a tela e abrindo na resolução padrão
telaPrincipal.geometry("%dx%d+%d+%d" % (width, height, posX, posY))
#Definindo um ícono para a aplicação
#Definindo a tela para abrir maximizada


abasFuncionais = ttk.Notebook(telaPrincipal, bootstyle="info")
frame1 = ttk.Frame(telaPrincipal)
frame2 = ttk.Frame(telaPrincipal)
frame3 = ttk.Frame(telaPrincipal)
frame4 = ttk.Frame(telaPrincipal)
frame5 = ttk.Frame(telaPrincipal)

frame1.pack(fill='both', expand=True)
frame2.pack(fill='both', expand=True)
frame3.pack(fill='both', expand=True)
frame4.pack(fill='both', expand=True)
frame5.pack(fill='both', expand=True)


abasFuncionais.add(frame1, text='Importador')
abasFuncionais.add(frame2, text='SQLite')
abasFuncionais.add(frame3, text='MySQL')
abasFuncionais.add(frame4, text='SQLServer')
abasFuncionais.add(frame5, text='Instruções')

abasFuncionais.place(relx=0.01, rely=0.015, relwidth=0.98, relheight=0.975)


abasFuncionais.tab(1, state="disabled")

abasFuncionais.select(4)

from tTelas.abaMySQL import *
from tTelas.abaSQLServer import *
from tTelas.abaInstrucoes import *
##Definindo as funcionalidades nas abas

#Itens no Frame 01 Importador
#Por preferência, esta aba não ficará seperada em um outro arquivo como as demais
# Obtém o caminho para o diretório do executável
diretorioExecutavel = os.path.dirname(os.path.abspath(__file__))

# Caminho do arquivo Excel de entrada
arquivoExcelConsumir = ''

# Nome e caminho do arquivo CSV de saída
arquivoCSVGerado = os.path.join(diretorioExecutavel, r"..\cGerado\arquivoDestino.csv")

# Nome do arquivo do banco de dados SQLite
arquivoBDSQLite = os.path.join(diretorioExecutavel, r"..\cGerado\exemploDados.db")

# Nome do arquivo do banco de dados MySQL
hostMySQL = ''
usuarioMySQL = ''
senhaMySQL = ''
bancoMySQL = ''
nomeTabela = ''
arquivoExcelConsumir = ''

hostSQLServer = ''
bancoSQLServer = ''

#Iniciando váriaveis para validações de comandos
varInserirDadosBDSQLite = ttk.StringVar()
varInserirDadosBDSQLite.set('0')

varInserirDadosBDMySQL = ttk.StringVar()
varInserirDadosBDMySQL.set('0')

varInserirDadosBDSQLServer = ttk.StringVar()
varInserirDadosBDSQLServer.set('0')


# Função para converter um arquivo Excel (xlsx) em CSV
def excelParaCsv(arquivoExcelConsumir, arquivoCSVGerado):
    global nomeTabela

    try:
        # Lê o arquivo Excel
        arquivo = pd.read_excel(arquivoExcelConsumir)
        
        nomeAbaExcel = load_workbook(filename=arquivoExcelConsumir, read_only=True)

        nomeTabela = nomeAbaExcel.sheetnames[0]

        print(nomeTabela)

        # Salva os dados em um arquivo CSV, tratando para aceitar possíveis caracteres especiais
        arquivo.to_csv(arquivoCSVGerado, index=False, encoding='utf-8-sig')
        
        notify("Status da Conversão", f"Conversão concluída. Arquivo CSV salvo como {arquivoCSVGerado}")
        print(f"Conversão concluída. Arquivo CSV salvo como {arquivoCSVGerado}")
    
    except Exception as e:
        print(f"Ocorreu um erro: {str(e)}")

# Função para ler o arquivo CSV e inserir os dados em uma tabela SQLite
def inserirDadosBDSQLite(arquivoCSVGerado, arquivoBDSQLite, nomeTabela):
    try:
        # Lê o arquivo CSV
        arquivo = pd.read_csv(arquivoCSVGerado, encoding='utf-8')

        # Cria uma conexão com o banco de dados SQLite
        conn = sqlite3.connect(arquivoBDSQLite)

        # Insere os dados no banco de dados SQLite
        arquivo.to_sql(nomeTabela, conn, if_exists='replace', index=False)

        notify("BD Local atualizado", f"Dados inseridos com sucesso na tabela '{nomeTabela}' do banco de dados '{arquivoBDSQLite}'")
        print(f"Dados inseridos com sucesso na tabela '{nomeTabela}' do banco de dados '{arquivoBDSQLite}'")

        # Fecha a conexão com o banco de dados
        conn.close()

    except Exception as e:
        print(f"Ocorreu um erro: {str(e)}")

# Função para ler o arquivo CSV e inserir os dados em uma tabela MySQL
def inserirDadosBDMySQL(arquivoCSVGerado, hostMySQL, usuarioMySQL, senhaMySQL, bancoMySQL, nomeTabela):       
    try:
        # Lê o arquivo CSV
        arquivo = pd.read_csv(arquivoCSVGerado, encoding='utf-8')

        # Cria uma conexão com o banco de dados MySQL
        conn = mysql.connector.connect(
            host=hostMySQL,
            user=usuarioMySQL,
            password=senhaMySQL,
            database=bancoMySQL
        )

        cursor = conn.cursor()

        # Cria a tabela no banco de dados MySQL (se não existir)
        create_table_query = f"CREATE TABLE IF NOT EXISTS {nomeTabela} ("
        for column in arquivo.columns:
            create_table_query += f"{column} VARCHAR(255), "
        create_table_query = create_table_query.rstrip(", ") + ")"
        cursor.execute(create_table_query)

        # Insere os dados na tabela MySQL
        for index, row in arquivo.iterrows():
            insert_query = f"INSERT INTO {nomeTabela} ({', '.join(arquivo.columns)}) VALUES ({', '.join(['%s']*len(arquivo.columns))})"
            cursor.execute(insert_query, tuple(row))

        conn.commit()
        notify("BD MySQL atualizado", f"Dados inseridos com sucesso na tabela '{nomeTabela}' do banco de dados MySQL")
        print(f"Dados inseridos com sucesso na tabela '{nomeTabela}' do banco de dados MySQL")

        # Fecha a conexão com o banco de dados MySQL
        cursor.close()
        conn.close()

    except Exception as e:
        print(f"Ocorreu um erro: {str(e)}")

def inserirDadosBDSQLServer(arquivoCSVGerado, nomeTabela):
    try:
        # Lê o arquivo CSV
        arquivo = pd.read_csv(arquivoCSVGerado, encoding='utf-8')

        # Cria uma conexão com o banco de dados SQL Server
        conn = pyodbc.connect("DRIVER={SQL Server};SERVER="+hostSQLServer+";DATABASE="+bancoSQLServer+";Trusted_Connection=yes")

        cursor = conn.cursor()

        # Cria a tabela no banco de dados SQL Server (se não existir)
        create_table_query = f"IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='{nomeTabela}' AND xtype='U') CREATE TABLE {nomeTabela} ("
        for column in arquivo.columns:
            create_table_query += f"{column} NVARCHAR(255), "
        create_table_query = create_table_query.rstrip(", ") + ")"
        cursor.execute(create_table_query)

        # Insere os dados na tabela SQL Server
        for index, row in arquivo.iterrows():
            insert_query = f"INSERT INTO {nomeTabela} ({', '.join(arquivo.columns)}) VALUES ({', '.join(['?']*len(arquivo.columns))})"
            cursor.execute(insert_query, tuple(row))

        conn.commit()
        notify("BD SQL Server atualizado", f"Dados inseridos com sucesso na tabela '{nomeTabela}' do banco de dados SQL Server")
        print(f"Dados inseridos com sucesso na tabela '{nomeTabela}' do banco de dados SQL Server")

        # Fecha a conexão com o banco de dados SQL Server
        cursor.close()
        conn.close()

    except Exception as e:
        print(f"Ocorreu um erro: {str(e)}")

labelDescricao = ttk.Label(frame1, text="Importar dados para:")
labelDescricao.place(relx=0.32, rely=0.12)

def validaAlteracaoBDSQLite():       
    if varInserirDadosBDSQLite.get() == '1':
        #print(varInserirDadosBDSQLite.get())
        pass

def validaAlteracaoBDMySQL():
    if varInserirDadosBDMySQL.get() == '1':
        global hostMySQL, usuarioMySQL, senhaMySQL, bancoMySQL

        abasFuncionais.select(2)

        hostMySQL = tituloServidorMySQL.get()
        usuarioMySQL = tituloUsuarioMySQL.get()
        senhaMySQL = tituloSenhaMySQL.get()
        bancoMySQL = tituloDatabaseMySQL.get()

def validaAlteracaoBDSQLServer():        
    if varInserirDadosBDSQLServer.get() == '1':
        global hostSQLServer, bancoSQLServer
        abasFuncionais.select(3)

        hostSQLServer = tituloServidorSQLServer.get()
        bancoSQLServer = tituloDatabaseSQLServer.get()
    
validaInserirDadosBDSQLite = ttk.Checkbutton(  frame1
                                   , text=None
                                   , command=validaAlteracaoBDSQLite
                                   , variable=varInserirDadosBDSQLite
                                   , onvalue=1
                                   , offvalue=0
                                   , bootstyle="round-toggle"
                                   )
validaInserirDadosBDSQLite.place(relx=0.53, rely=0.25)

labelInserirDadosBDSQLServer = ttk.Label(frame1, text="SQLite")
labelInserirDadosBDSQLServer.place(relx=0.33, rely=0.24)

validaInserirDadosBDMySQL = ttk.Checkbutton(  frame1
                                   , text=None
                                   , command=validaAlteracaoBDMySQL
                                   , variable=varInserirDadosBDMySQL
                                   , onvalue=1
                                   , offvalue=0
                                   , bootstyle="round-toggle"
                                   )
validaInserirDadosBDMySQL.place(relx=0.53, rely=0.35)

labelInserirDadosBDSQLServer = ttk.Label(frame1, text="MySQL")
labelInserirDadosBDSQLServer.place(relx=0.33, rely=0.34)

validaInserirDadosBDSQLServer = ttk.Checkbutton(  frame1
                                   , text=None
                                   , command=validaAlteracaoBDSQLServer
                                   , variable=varInserirDadosBDSQLServer
                                   , onvalue=1
                                   , offvalue=0
                                   , bootstyle="round-toggle"
                                   )
validaInserirDadosBDSQLServer.place(relx=0.53, rely=0.45)

labelInserirDadosBDSQLServer = ttk.Label(frame1, text="SQL Server")
labelInserirDadosBDSQLServer.place(relx=0.33, rely=0.44)




def selecionarArquivoExcel():
    global arquivoExcelConsumir
    arquivoExcelConsumir = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx")])
    if arquivoExcelConsumir:
        tituloArquivoSelecionado.configure(state="active")
        tituloArquivoSelecionado.insert(0, arquivoExcelConsumir)
        tituloArquivoSelecionado.configure(state="readonly")


def processarDados():
    global arquivoExcelConsumir
    if arquivoExcelConsumir == '':
        messagebox.showwarning("Aviso", "É necessário selecionar o arquivo Excel antes iniciar o processamento")

    if varInserirDadosBDMySQL.get() == '1':
        if tituloServidorMySQL.get() == '' or tituloUsuarioMySQL.get() == '' or tituloSenhaMySQL.get() == '' or tituloDatabaseMySQL.get() == '':
            messagebox.showwarning("Aviso", "É necessário preencher todos os campos da aba MySQL")
            abasFuncionais.select(2)
            tituloArquivoSelecionado.configure(state="active")
            tituloArquivoSelecionado.delete(0, END)
            arquivoExcelConsumir = ''
            tituloArquivoSelecionado.configure(state="readonly")

    if varInserirDadosBDSQLServer.get() == '1':
        if tituloServidorSQLServer.get() == '' or tituloDatabaseSQLServer.get() == '':
            messagebox.showwarning("Aviso", "É necessário preencher todos os campos da aba SQL Server")
            abasFuncionais.select(3)
            tituloArquivoSelecionado.configure(state="active")
            tituloArquivoSelecionado.delete(0, END)
            arquivoExcelConsumir = ''
            tituloArquivoSelecionado.configure(state="readonly")
    
    if arquivoExcelConsumir:
        try:
            # Chama a função para converter o arquivo
            excelParaCsv(arquivoExcelConsumir, arquivoCSVGerado)

            if varInserirDadosBDSQLite.get() == '1':
                # Chama a função para inserir os dados no SQLite
                inserirDadosBDSQLite(arquivoCSVGerado, arquivoBDSQLite, nomeTabela)

            if varInserirDadosBDMySQL.get() == '1':
                inserirDadosBDMySQL(arquivoCSVGerado, hostMySQL, usuarioMySQL, senhaMySQL, bancoMySQL, nomeTabela)

            if varInserirDadosBDSQLServer.get() == '1':
                # Chama a função para inserir os dados no SQlServer
                inserirDadosBDSQLServer(arquivoCSVGerado, nomeTabela)

            messagebox.showinfo("Sucesso", "O processo de importação de dados foi realizado com sucesso...")
        except Exception as e:
            messagebox.showerror("Erro!!!",f"Ocorreu um erro: {str(e)}")

selecionarArquivo = ttk.Button(frame1, text="Selecionar Arquivo Excel", command=selecionarArquivoExcel)
selecionarArquivo.place(relx=0.1, rely=0.6)

processarUpload = ttk.Button(frame1, text="Processar Dados", command=processarDados)
processarUpload.place(relx=0.6, rely=0.6)


labelArquivoSelecionado = ttk.Label(frame1, text="Arquivo Selecionado", font=ctk.CTkFont(size=10, weight="bold"))
labelArquivoSelecionado.place(relx=0.35, rely=0.73)
tituloArquivoSelecionado = ttk.Entry(frame1, state="readonly")
tituloArquivoSelecionado.place(relx=0.01, rely=0.8, relwidth=0.98, relheight=0.13)


###################################
telaPrincipal.mainloop()
